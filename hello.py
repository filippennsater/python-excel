from openpyxl.workbook import Workbook
from openpyxl import load_workbook

#create a workbbok object
#wb = Workbook()



#load existing spreadsheet
wb = load_workbook('hello.xlsx')

#create an active worksheet
ws = wb.active

#print something from the spreadsheet
'''
print("This is the result")
print(ws["A2"])

print(ws["A2"].value)

print(f'{ws["A2"].value}: {ws["B2"].value}')

name = ws["A2"].value

color = ws["B2"].value

print(f'{name}: {color}')

#grab a a whole column

column_a = ws['A']
print(column_a)

for cell in column_a:
    print(f'{cell.value}\n')
'''

#grab a range

rows = ws['2':'10']



for x in rows:
    for y in x:
        print(y.value)